"""
Excel 내보내기 뷰
"""
from datetime import datetime
from io import BytesIO

from django.shortcuts import render
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from students.models import Student
from attendance.models import Attendance
from payments.models import Payment
from classes.models import Class


def create_excel_response(workbook, filename):
    """Excel 파일 HTTP 응답 생성"""
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def style_header(ws, row=1):
    """헤더 스타일 적용"""
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in ws[row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border


@login_required
def export_hub(request):
    """내보내기 허브 페이지"""
    classes = Class.objects.filter(is_active=True)
    return render(request, 'exports/hub.html', {
        'classes': classes,
    })


@login_required
def export_students(request):
    """학생 목록 Excel 내보내기"""
    # 필터
    class_id = request.GET.get('class_id')
    status = request.GET.get('status')
    
    students = Student.objects.select_related('assigned_class').order_by('name')
    
    if class_id:
        students = students.filter(assigned_class_id=class_id)
    if status:
        students = students.filter(status=status)
    
    # Excel 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "학생 목록"
    
    # 헤더
    headers = ['이름', '성별', '생년월일', '연락처', '학부모 이름', '학부모 연락처', 
               '반', '재원 상태', '등록일', '학교', '학년']
    ws.append(headers)
    style_header(ws)
    
    # 데이터
    status_map = dict(Student.STATUS_CHOICES)
    gender_map = dict(Student.GENDER_CHOICES)
    
    for student in students:
        ws.append([
            student.name,
            gender_map.get(student.gender, ''),
            student.birth_date.strftime('%Y-%m-%d') if student.birth_date else '',
            student.phone,
            student.parent_name,
            student.parent_phone,
            student.assigned_class.name if student.assigned_class else '',
            status_map.get(student.status, student.status),
            student.enrollment_date.strftime('%Y-%m-%d') if student.enrollment_date else '',
            student.school_name,
            student.grade,
        ])
    
    # 열 너비 조정
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 30)
    
    filename = f"students_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return create_excel_response(wb, filename)


@login_required
def export_attendance(request):
    """출결 Excel 내보내기"""
    # 필터
    class_id = request.GET.get('class_id')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    status = request.GET.get('status')
    
    attendances = Attendance.objects.select_related('student', 'assigned_class').order_by('-date', 'student__name')
    
    if class_id:
        attendances = attendances.filter(assigned_class_id=class_id)
    if date_from:
        attendances = attendances.filter(date__gte=date_from)
    if date_to:
        attendances = attendances.filter(date__lte=date_to)
    if status:
        attendances = attendances.filter(status=status)
    
    # Excel 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "출결 기록"
    
    # 헤더
    headers = ['날짜', '학생', '반', '상태', '메모']
    ws.append(headers)
    style_header(ws)
    
    # 데이터
    status_map = dict(Attendance.STATUS_CHOICES)
    
    for att in attendances:
        ws.append([
            att.date.strftime('%Y-%m-%d'),
            att.student.name,
            att.assigned_class.name if att.assigned_class else '',
            status_map.get(att.status, att.status),
            att.note,
        ])
    
    # 열 너비 조정
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 30)
    
    filename = f"attendance_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return create_excel_response(wb, filename)


@login_required
def export_payments(request):
    """수납 Excel 내보내기"""
    # 필터
    class_id = request.GET.get('class_id')
    year = request.GET.get('year')
    month = request.GET.get('month')
    status = request.GET.get('status')
    
    payments = Payment.objects.select_related('student', 'student__assigned_class').order_by('-year', '-month', 'student__name')
    
    if class_id:
        payments = payments.filter(student__assigned_class_id=class_id)
    if year:
        payments = payments.filter(year=year)
    if month:
        payments = payments.filter(month=month)
    if status:
        payments = payments.filter(status=status)
    
    # Excel 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "수납 내역"
    
    # 헤더
    headers = ['년도', '월', '학생', '반', '청구금액', '납부금액', '미납금액', '상태', '결제방식', '납부일', '메모']
    ws.append(headers)
    style_header(ws)
    
    # 데이터
    status_map = dict(Payment.STATUS_CHOICES)
    method_map = dict(Payment.PAYMENT_METHOD_CHOICES)
    
    for payment in payments:
        ws.append([
            payment.year,
            payment.month,
            payment.student.name,
            payment.student.assigned_class.name if payment.student.assigned_class else '',
            payment.amount,
            payment.paid_amount,
            payment.remaining_amount,
            status_map.get(payment.status, payment.status),
            method_map.get(payment.payment_method, payment.payment_method),
            payment.payment_date.strftime('%Y-%m-%d') if payment.payment_date else '',
            payment.note,
        ])
    
    # 열 너비 조정
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 30)
    
    filename = f"payments_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return create_excel_response(wb, filename)
